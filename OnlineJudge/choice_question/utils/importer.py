from django.db import transaction
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
import json
import re
from typing import List, Dict, Any, Optional
from ..models import ChoiceQuestion, Category, QuestionTag, ExamPaper, ExamPaperQuestion
from .validator import QuestionValidator
from .helpers import generate_question_id, format_answer


class QuestionImporter:
    """
    选择题导入工具类
    支持Excel、JSON等格式的题目批量导入
    """
    
    def __init__(self, user=None):
        self.user = user
        self.validator = QuestionValidator()
        self.import_log = []
        self.success_count = 0
        self.error_count = 0
    
    def import_from_excel(self, file_path: str, sheet_name: str = None) -> Dict[str, Any]:
        """
        从Excel文件导入选择题
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，默认使用第一个工作表
            
        Returns:
            导入结果统计
        """
        try:
            workbook = load_workbook(file_path, read_only=True)
            if sheet_name:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # 读取表头
            headers = [cell.value for cell in worksheet[1]]
            header_map = self._create_header_map(headers)
            
            questions_data = []
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # 跳过空行
                    continue
                    
                question_data = self._parse_excel_row(row, header_map, row_num)
                if question_data:
                    questions_data.append(question_data)
            
            return self._batch_import_questions(questions_data)
            
        except Exception as e:
            self._add_log('error', f'Excel文件读取失败: {str(e)}')
            return self._get_import_result()
    
    def import_from_json(self, file_path: str) -> Dict[str, Any]:
        """
        从JSON文件导入选择题
        
        Args:
            file_path: JSON文件路径
            
        Returns:
            导入结果统计
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if isinstance(data, list):
                questions_data = data
            elif isinstance(data, dict) and 'questions' in data:
                questions_data = data['questions']
            else:
                raise ValueError('JSON格式不正确，应包含questions数组')
            
            # 转换数据格式
            converted_questions = []
            for i, question in enumerate(questions_data):
                converted_question = self._convert_json_question_format(question, i + 1)
                converted_questions.append(converted_question)
            
            return self._batch_import_questions(converted_questions)
            
        except Exception as e:
            self._add_log('error', f'JSON文件读取失败: {str(e)}')
            return self._get_import_result()
    
    def _create_header_map(self, headers: List[str]) -> Dict[str, int]:
        """
        创建表头映射
        """
        header_map = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            header = str(header).strip().lower()
            
            # 标题映射
            if header in ['题目', '标题', 'title', 'question']:
                header_map['title'] = i
            elif header in ['描述', '题目描述', 'description', 'content']:
                header_map['description'] = i
            elif header in ['类型', '题目类型', 'type', 'question_type']:
                header_map['question_type'] = i
            elif header in ['选项', '选项列表', 'options', 'choices']:
                header_map['options'] = i
            elif header in ['答案', '正确答案', 'answer', 'correct_answer']:
                header_map['answer'] = i
            elif header in ['解析', '答案解析', 'explanation', 'analysis']:
                header_map['explanation'] = i
            elif header in ['难度', 'difficulty', 'level']:
                header_map['difficulty'] = i
            elif header in ['分值', '分数', 'score', 'points']:
                header_map['score'] = i
            elif header in ['分类', '类别', 'category']:
                header_map['category'] = i
            elif header in ['标签', 'tags', 'labels']:
                header_map['tags'] = i
        
        return header_map
    
    def _parse_excel_row(self, row: tuple, header_map: Dict[str, int], row_num: int) -> Optional[Dict[str, Any]]:
        """
        解析Excel行数据
        """
        try:
            question_data = {
                'row_num': row_num,
                'title': self._get_cell_value(row, header_map.get('title')),
                'description': self._get_cell_value(row, header_map.get('description', '')) or '',
                'question_type': self._parse_question_type(self._get_cell_value(row, header_map.get('question_type'))),
                'options': self._parse_options(self._get_cell_value(row, header_map.get('options'))),
                'answer': self._get_cell_value(row, header_map.get('answer')),
                'explanation': self._get_cell_value(row, header_map.get('explanation', '')) or '',
                'difficulty': self._parse_difficulty(self._get_cell_value(row, header_map.get('difficulty'))),
                'score': self._parse_score(self._get_cell_value(row, header_map.get('score'))),
                'category': self._get_cell_value(row, header_map.get('category', '')) or '',
                'tags': self._parse_tags(self._get_cell_value(row, header_map.get('tags', '')) or ''),
            }
            
            # 验证必填字段
            if not question_data['title']:
                self._add_log('error', f'第{row_num}行：题目标题不能为空')
                return None
            
            if not question_data['options']:
                self._add_log('error', f'第{row_num}行：选项不能为空')
                return None
            
            if not question_data['answer']:
                self._add_log('error', f'第{row_num}行：答案不能为空')
                return None
            
            return question_data
            
        except Exception as e:
            self._add_log('error', f'第{row_num}行数据解析失败: {str(e)}')
            return None
    
    def _get_cell_value(self, row: tuple, index: Optional[int]) -> str:
        """
        获取单元格值
        """
        if index is None or index >= len(row):
            return ''
        value = row[index]
        return str(value).strip() if value is not None else ''
    
    def _parse_question_type(self, value: str) -> str:
        """
        解析题目类型
        """
        if not value:
            return 'single'
        
        value = value.lower()
        if value in ['单选', 'single', '单选题', 'single_choice']:
            return 'single'
        elif value in ['多选', 'multiple', '多选题', 'multiple_choice']:
            return 'multiple'
        else:
            return 'single'  # 默认单选
    
    def _parse_options(self, value: str) -> List[str]:
        """
        解析选项
        支持多种格式：
        - A.选项1\nB.选项2\nC.选项3
        - A:选项1;B:选项2;C:选项3
        - ["选项1", "选项2", "选项3"]
        """
        if not value:
            return []
        
        try:
            # 尝试JSON格式
            if value.strip().startswith('['):
                return json.loads(value)
            
            # 按行分割
            if '\n' in value:
                lines = value.split('\n')
                options = []
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    # 移除选项标号 (A. B. C. 等)
                    line = re.sub(r'^[A-Z][.:]\s*', '', line, flags=re.IGNORECASE)
                    options.append(line)
                return options
            
            # 按分号分割
            if ';' in value:
                options = []
                for option in value.split(';'):
                    option = option.strip()
                    if not option:
                        continue
                    # 移除选项标号
                    option = re.sub(r'^[A-Z][.:]\s*', '', option, flags=re.IGNORECASE)
                    options.append(option)
                return options
            
            # 单个选项
            return [value.strip()]
            
        except Exception:
            return [value.strip()] if value.strip() else []
    
    def _parse_difficulty(self, value: str) -> str:
        """
        解析难度
        """
        if not value:
            return 'medium'
        
        value = value.lower()
        if value in ['简单', 'easy', '容易', '1']:
            return 'easy'
        elif value in ['中等', 'medium', '普通', '2']:
            return 'medium'
        elif value in ['困难', 'hard', '难', '3']:
            return 'hard'
        else:
            return 'medium'
    
    def _parse_score(self, value: str) -> int:
        """
        解析分值
        """
        if not value:
            return 1
        
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 1
    
    def _parse_tags(self, value: str) -> List[str]:
        """
        解析标签
        """
        if not value:
            return []
        
        # 支持逗号、分号、空格分割
        tags = re.split(r'[,;，；\s]+', value)
        return [tag.strip() for tag in tags if tag.strip()]
    
    def _batch_import_questions(self, questions_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        批量导入题目
        """
        with transaction.atomic():
            for i, question_data in enumerate(questions_data):
                try:
                    # 添加行号用于错误提示，但不包含在验证数据中
                    row_num = i + 1
                    question_data['row_num'] = row_num  # 使用row_num而不是_row_num
                    self._import_single_question(question_data)
                except Exception as e:
                    self._add_log('error', f'第{row_num}行导入失败: {str(e)}')
        
        return self._get_import_result()
    
    def _convert_json_question_format(self, question_data: Dict[str, Any], row_num: int) -> Dict[str, Any]:
        """
        转换JSON题目格式到系统内部格式
        """
        converted = {}
        
        # 转换题目类型
        if 'type' in question_data:
            question_type = question_data['type'].lower()
            converted['question_type'] = 'single' if question_type == 'single' else 'multiple'
        else:
            converted['question_type'] = 'single'
        
        # 转换题目内容
        if 'question' in question_data:
            # 清理标题中的特殊字符
            title = question_data['question']
            # 移除或替换特殊字符
            title = title.replace('<', '＜').replace('>', '＞').replace('"', '"').replace('/', '／').replace('\\', '＼')
            converted['title'] = title[:200]  # 限制标题长度
            converted['description'] = question_data['question']  # 完整描述
        else:
            converted['title'] = '未知题目'
            converted['description'] = ''
        
        # 转换选项
        converted['options'] = question_data.get('options', [])
        
        # 转换答案
        if 'correct' in question_data:
            correct_answer = question_data['correct']
            if isinstance(correct_answer, str):
                # 如果是字母答案（如"A", "B"），转换为索引字符串
                if correct_answer.upper() in ['A', 'B', 'C', 'D', 'E', 'F']:
                    answer_index = ord(correct_answer.upper()) - ord('A')
                    converted['answer'] = str(answer_index)
                else:
                    # 如果是数字字符串，直接使用
                    converted['answer'] = str(correct_answer)
            else:
                converted['answer'] = str(correct_answer)
        else:
            converted['answer'] = '0'
        
        # 转换解析
        converted['explanation'] = question_data.get('explanation', '')
        
        # 设置默认值
        converted['difficulty'] = 'medium'
        converted['score'] = 1
        
        # 如果有ID字段，可以作为分类使用
        if 'id' in question_data:
            # 从ID中提取分类信息（如GESP_2_2024_3_1中的GESP）
            id_parts = question_data['id'].split('_')
            if len(id_parts) > 0:
                converted['category'] = id_parts[0]
        else:
            converted['category'] = ''
        
        converted['tags'] = []
        
        return converted
    
    def _import_single_question(self, question_data: Dict[str, Any]):
        """
        导入单个题目
        """
        # 验证数据
        validation_result = self.validator.validate_question_data(question_data)
        if not validation_result['is_valid']:
            raise ValidationError('; '.join(validation_result['errors']))
        
        # 处理分类
        category = None
        if question_data.get('category'):
            category, _ = Category.objects.get_or_create(
                name=question_data['category'],
                defaults={'description': f'自动创建的分类: {question_data["category"]}'}
            )
        
        # 创建题目
        question = ChoiceQuestion.objects.create(
            _id=generate_question_id(),
            title=question_data['title'],
            description=question_data['description'],
            question_type=question_data['question_type'],
            options=json.dumps(question_data['options'], ensure_ascii=False),
            correct_answer=question_data.get('correct', 'A'),
            explanation=question_data['explanation'],
            difficulty=question_data['difficulty'],
            score=question_data['score'],
            category=category,
            created_by=self.user,
            is_public=True
        )
        
        # 处理标签
        if question_data.get('tags'):
            for tag_name in question_data['tags']:
                tag, _ = QuestionTag.objects.get_or_create(
                    name=tag_name,
                    defaults={'description': f'自动创建的标签: {tag_name}'}
                )
                question.tags.add(tag)
        
        self.success_count += 1
        row_num = question_data.get('row_num', '未知')
        self._add_log('success', f'第{row_num}行导入成功: {question.title}')
    
    def _add_log(self, level: str, message: str):
        """
        添加日志
        """
        self.import_log.append({
            'level': level,
            'message': message
        })
        
        if level == 'error':
            self.error_count += 1
    
    def _get_import_result(self) -> Dict[str, Any]:
        """
        获取导入结果
        """
        return {
            'success_count': self.success_count,
            'error_count': self.error_count,
            'total_count': self.success_count + self.error_count,
            'logs': self.import_log
        }
    
    def import_paper_from_json(self, file_path: str, category_name: str = None, paper_title: str = None, language: str = None, use_import_order: bool = True) -> Dict[str, Any]:
        """
        从JSON文件导入整套试卷
        
        Args:
            file_path: JSON文件路径
            category_name: 分类名称
            paper_title: 试卷标题
            language: 编程语言
            use_import_order: 是否按导入顺序排序
            
        Returns:
            导入结果，包含试卷ID和统计信息
        """
        with transaction.atomic():
            # 1. 解析JSON文件
            paper_data = self._parse_paper_json(file_path)
            questions_data = paper_data.get('questions', [])
            
            # 2. 创建或获取分类
            category = None
            if category_name:
                category, _ = Category.objects.get_or_create(
                    name=category_name,
                    defaults={'description': f'自动创建的分类: {category_name}'}
                )
            
            # 3. 批量导入题目，保持顺序
            imported_questions = []
            for i, question_data in enumerate(questions_data):
                question_data['import_order'] = i + 1  # 设置导入顺序
                if category:
                    question_data['category'] = category.name
                if language:
                    question_data['language'] = language
                question = self._import_single_question_for_paper(question_data)
                imported_questions.append(question)
            
            # 4. 创建固定题目试卷
            paper = ExamPaper.objects.create(
                title=paper_title or paper_data.get('title', '导入的试卷'),
                description=paper_data.get('description', ''),
                duration=paper_data.get('duration', 60),
                total_score=sum(q.score for q in imported_questions),
                question_count=len(imported_questions),
                paper_type='fixed',
                use_import_order=use_import_order,
                is_active=True,
                created_by=self.user
            )
            
            # 5. 建立试卷题目关联
            for i, question in enumerate(imported_questions):
                ExamPaperQuestion.objects.create(
                    paper=paper,
                    question=question,
                    order=i + 1,
                    score=question.score
                )
            
            # 6. 关联分类
            if category:
                paper.categories.add(category)
            
            return {
                'paper_id': paper.id,
                'paper_title': paper.title,
                'success_count': len(imported_questions),
                'error_count': 0,
                'total_count': len(imported_questions),
                'logs': self.import_log
            }
    
    def _parse_paper_json(self, file_path: str) -> Dict[str, Any]:
        """
        解析试卷JSON文件
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 验证JSON格式
        if 'questions' not in data:
            raise ValidationError('JSON文件必须包含questions字段')
        
        if not isinstance(data['questions'], list):
            raise ValidationError('questions字段必须是数组')
        
        return data
    
    def _import_single_question_for_paper(self, question_data: Dict[str, Any]) -> ChoiceQuestion:
        """
        为试卷导入单个题目（返回题目对象）
        """
        # 验证数据
        validation_result = self.validator.validate_question_data(question_data)
        if not validation_result['is_valid']:
            raise ValidationError('; '.join(validation_result['errors']))
        
        # 处理分类
        category = None
        if question_data.get('category'):
            category, _ = Category.objects.get_or_create(
                name=question_data['category'],
                defaults={'description': f'自动创建的分类: {question_data["category"]}'}
            )
        
        # 创建题目
        question = ChoiceQuestion.objects.create(
            _id=generate_question_id(),
            title=question_data['title'],
            description=question_data['description'],
            question_type=question_data['question_type'],
            options=question_data['options'],  # 直接传递对象数组，让JSONField自动处理序列化
            correct_answer=question_data.get('correct', 'A'),
            explanation=question_data['explanation'],
            difficulty=question_data['difficulty'],
            score=question_data['score'],
            import_order=question_data.get('import_order', 0),
            language=question_data.get('language'),
            category=category,
            created_by=self.user,
            is_public=True
        )
        
        # 处理标签
        if question_data.get('tags'):
            for tag_name in question_data['tags']:
                tag, _ = QuestionTag.objects.get_or_create(
                    name=tag_name,
                    defaults={'description': f'自动创建的标签: {tag_name}'}
                )
                question.tags.add(tag)
        
        self.success_count += 1
        self._add_log('success', f'题目导入成功: {question.title}')
        
        return question