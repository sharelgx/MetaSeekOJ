from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json
import csv
import io
from typing import List, Dict, Any, Optional
from django.db.models import QuerySet
from ..models import ChoiceQuestion


class QuestionExporter:
    """
    选择题导出工具类
    支持Excel、JSON、CSV等格式的题目批量导出
    """
    
    def __init__(self):
        self.export_fields = [
            'question_id', 'title', 'description', 'question_type',
            'options', 'answer', 'explanation', 'difficulty', 'score',
            'category', 'tags', 'total_submissions', 'correct_submissions',
            'accuracy_rate', 'created_by', 'create_time', 'is_public'
        ]
    
    def export_to_excel(self, questions: QuerySet, filename: str = None) -> HttpResponse:
        """
        导出为Excel格式
        
        Args:
            questions: 题目查询集
            filename: 文件名
            
        Returns:
            HttpResponse对象
        """
        if not filename:
            filename = 'choice_questions.xlsx'
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '选择题列表'
        
        # 设置表头
        headers = [
            '题目ID', '标题', '描述', '类型', '选项', '答案', '解析',
            '难度', '分值', '分类', '标签', '提交次数', '正确次数',
            '正确率', '创建者', '创建时间', '是否公开'
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row, question in enumerate(questions, 2):
            question_data = self._serialize_question(question)
            
            worksheet.cell(row=row, column=1, value=question_data['question_id'])
            worksheet.cell(row=row, column=2, value=question_data['title'])
            worksheet.cell(row=row, column=3, value=question_data['description'])
            worksheet.cell(row=row, column=4, value=question_data['question_type_display'])
            worksheet.cell(row=row, column=5, value=question_data['options_text'])
            worksheet.cell(row=row, column=6, value=question_data['answer_text'])
            worksheet.cell(row=row, column=7, value=question_data['explanation'])
            worksheet.cell(row=row, column=8, value=question_data['difficulty_display'])
            worksheet.cell(row=row, column=9, value=question_data['score'])
            worksheet.cell(row=row, column=10, value=question_data['category_name'])
            worksheet.cell(row=row, column=11, value=question_data['tags_text'])
            worksheet.cell(row=row, column=12, value=question_data['total_submissions'])
            worksheet.cell(row=row, column=13, value=question_data['correct_submissions'])
            worksheet.cell(row=row, column=14, value=f"{question_data['accuracy_rate']:.1f}%")
            worksheet.cell(row=row, column=15, value=question_data['created_by'])
            worksheet.cell(row=row, column=16, value=question_data['create_time'])
            worksheet.cell(row=row, column=17, value='是' if question_data['is_public'] else '否')
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                cell_value = str(worksheet[f'{column_letter}{row}'].value or '')
                max_length = max(max_length, len(cell_value))
            
            # 设置列宽，最小10，最大50
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 保存到响应
        workbook.save(response)
        return response
    
    def export_to_json(self, questions: QuerySet, filename: str = None) -> HttpResponse:
        """
        导出为JSON格式
        
        Args:
            questions: 题目查询集
            filename: 文件名
            
        Returns:
            HttpResponse对象
        """
        if not filename:
            filename = 'choice_questions.json'
        
        questions_data = []
        for question in questions:
            question_data = self._serialize_question(question, for_json=True)
            questions_data.append(question_data)
        
        export_data = {
            'export_info': {
                'total_count': len(questions_data),
                'export_time': questions.first().create_time.isoformat() if questions_data else None,
                'format_version': '1.0'
            },
            'questions': questions_data
        }
        
        response = HttpResponse(
            json.dumps(export_data, ensure_ascii=False, indent=2),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def export_to_csv(self, questions: QuerySet, filename: str = None) -> HttpResponse:
        """
        导出为CSV格式
        
        Args:
            questions: 题目查询集
            filename: 文件名
            
        Returns:
            HttpResponse对象
        """
        if not filename:
            filename = 'choice_questions.csv'
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        headers = [
            '题目ID', '标题', '描述', '类型', '选项', '答案', '解析',
            '难度', '分值', '分类', '标签', '提交次数', '正确次数',
            '正确率', '创建者', '创建时间', '是否公开'
        ]
        writer.writerow(headers)
        
        # 写入数据
        for question in questions:
            question_data = self._serialize_question(question)
            row = [
                question_data['question_id'],
                question_data['title'],
                question_data['description'],
                question_data['question_type_display'],
                question_data['options_text'],
                question_data['answer_text'],
                question_data['explanation'],
                question_data['difficulty_display'],
                question_data['score'],
                question_data['category_name'],
                question_data['tags_text'],
                question_data['total_submissions'],
                question_data['correct_submissions'],
                f"{question_data['accuracy_rate']:.1f}%",
                question_data['created_by'],
                question_data['create_time'],
                '是' if question_data['is_public'] else '否'
            ]
            writer.writerow(row)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 添加BOM以支持Excel正确显示中文
        response.write('\ufeff')
        
        return response
    
    def export_template(self, format_type: str = 'excel') -> HttpResponse:
        """
        导出导入模板
        
        Args:
            format_type: 格式类型 ('excel', 'csv')
            
        Returns:
            HttpResponse对象
        """
        if format_type == 'excel':
            return self._create_excel_template()
        elif format_type == 'csv':
            return self._create_csv_template()
        else:
            raise ValueError(f'不支持的格式类型: {format_type}')
    
    def _create_excel_template(self) -> HttpResponse:
        """
        创建Excel导入模板
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '选择题导入模板'
        
        # 设置表头
        headers = [
            '题目标题*', '题目描述', '题目类型*', '选项*', '答案*',
            '解析', '难度', '分值', '分类', '标签'
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据
        example_data = [
            [
                'Python中哪个关键字用于定义函数？',
                '这是一道关于Python基础语法的题目',
                '单选',
                'A.def\nB.function\nC.define\nD.func',
                'A',
                'def是Python中定义函数的关键字',
                '简单',
                '2',
                'Python基础',
                'Python,语法,函数'
            ],
            [
                '以下哪些是Python的数据类型？',
                '选择所有正确的Python数据类型',
                '多选',
                'A.int\nB.string\nC.list\nD.dict',
                'A,C,D',
                'int、list、dict都是Python的内置数据类型，string应该是str',
                '中等',
                '3',
                'Python基础',
                'Python,数据类型'
            ]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row, column=col, value=value)
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 20
        
        # 添加说明工作表
        help_sheet = workbook.create_sheet('导入说明')
        help_content = [
            ['字段说明', ''],
            ['题目标题*', '必填，题目的标题'],
            ['题目描述', '可选，题目的详细描述'],
            ['题目类型*', '必填，单选 或 多选'],
            ['选项*', '必填，格式：A.选项1\\nB.选项2\\nC.选项3'],
            ['答案*', '必填，单选填A，多选填A,C'],
            ['解析', '可选，答案解析'],
            ['难度', '可选，简单/中等/困难，默认中等'],
            ['分值', '可选，数字，默认1'],
            ['分类', '可选，题目分类'],
            ['标签', '可选，多个标签用逗号分隔'],
            ['', ''],
            ['注意事项', ''],
            ['1. 带*的字段为必填项', ''],
            ['2. 选项格式支持多种方式：', ''],
            ['   - A.选项1\\nB.选项2', ''],
            ['   - A:选项1;B:选项2', ''],
            ['3. 答案格式：', ''],
            ['   - 单选：A 或 1', ''],
            ['   - 多选：A,C 或 1,3', ''],
        ]
        
        for row, (key, value) in enumerate(help_content, 1):
            help_sheet.cell(row=row, column=1, value=key)
            help_sheet.cell(row=row, column=2, value=value)
            if key in ['字段说明', '注意事项']:
                help_sheet.cell(row=row, column=1).font = Font(bold=True)
        
        help_sheet.column_dimensions['A'].width = 20
        help_sheet.column_dimensions['B'].width = 40
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="choice_questions_template.xlsx"'
        
        workbook.save(response)
        return response
    
    def _create_csv_template(self) -> HttpResponse:
        """
        创建CSV导入模板
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        headers = [
            '题目标题*', '题目描述', '题目类型*', '选项*', '答案*',
            '解析', '难度', '分值', '分类', '标签'
        ]
        writer.writerow(headers)
        
        # 添加示例数据
        example_data = [
            [
                'Python中哪个关键字用于定义函数？',
                '这是一道关于Python基础语法的题目',
                '单选',
                'A.def;B.function;C.define;D.func',
                'A',
                'def是Python中定义函数的关键字',
                '简单',
                '2',
                'Python基础',
                'Python,语法,函数'
            ]
        ]
        
        for data in example_data:
            writer.writerow(data)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = 'attachment; filename="choice_questions_template.csv"'
        
        # 添加BOM
        response.write('\ufeff')
        
        return response
    
    def _serialize_question(self, question: ChoiceQuestion, for_json: bool = False) -> Dict[str, Any]:
        """
        序列化题目数据
        
        Args:
            question: 题目对象
            for_json: 是否用于JSON导出
            
        Returns:
            序列化后的数据
        """
        # 解析选项
        try:
            options = json.loads(question.options) if question.options else []
        except (json.JSONDecodeError, TypeError):
            options = []
        
        # 格式化选项文本
        if for_json:
            options_text = options
        else:
            options_text = '\n'.join([f'{chr(65+i)}.{opt}' for i, opt in enumerate(options)])
        
        # 格式化答案
        try:
            answer_list = json.loads(question.answer) if question.answer else []
            if for_json:
                answer_text = answer_list
            else:
                if question.question_type == 'single':
                    answer_text = chr(65 + answer_list[0]) if answer_list else ''
                else:
                    answer_text = ','.join([chr(65 + ans) for ans in answer_list])
        except (json.JSONDecodeError, TypeError, IndexError):
            answer_text = question.answer or ''
        
        # 计算正确率
        accuracy_rate = 0
        if question.total_submissions > 0:
            accuracy_rate = (question.correct_submissions / question.total_submissions) * 100
        
        # 获取标签
        tags = list(question.tags.values_list('name', flat=True))
        tags_text = ','.join(tags) if not for_json else tags
        
        # 类型显示
        type_display_map = {
            'single': '单选',
            'multiple': '多选'
        }
        
        # 难度显示
        difficulty_display_map = {
            'easy': '简单',
            'medium': '中等',
            'hard': '困难'
        }
        
        data = {
            'question_id': question.question_id,
            'title': question.title,
            'description': question.description or '',
            'question_type': question.question_type,
            'question_type_display': type_display_map.get(question.question_type, question.question_type),
            'options': options if for_json else None,
            'options_text': options_text,
            'answer': answer_list if for_json else None,
            'answer_text': answer_text,
            'explanation': question.explanation or '',
            'difficulty': question.difficulty,
            'difficulty_display': difficulty_display_map.get(question.difficulty, question.difficulty),
            'score': question.score,
            'category_name': question.category.name if question.category else '',
            'tags': tags if for_json else None,
            'tags_text': tags_text,
            'total_submissions': question.total_submissions,
            'correct_submissions': question.correct_submissions,
            'accuracy_rate': accuracy_rate,
            'created_by': question.created_by.username if question.created_by else '',
            'create_time': question.create_time.strftime('%Y-%m-%d %H:%M:%S') if not for_json else question.create_time.isoformat(),
            'is_public': question.is_public
        }
        
        if for_json:
            # JSON格式包含更多字段
            data.update({
                'last_update_time': question.last_update_time.isoformat(),
                'plugin_version': question.plugin_version
            })
        
        return data